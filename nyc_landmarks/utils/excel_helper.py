import os
from typing import Any, Dict, List, Mapping, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


def export_dicts_to_excel(
    data: List[Dict[str, Any]],
    output_file_path: str,
    column_order: Optional[List[str]] = None,
    header_map: Optional[Dict[str, str]] = None,
) -> None:
    """
    Export a list of dictionaries to an Excel file.

    Args:
        data (list): List of dictionaries to export.
        output_file_path (str): Path to the Excel file to create.
        column_order (list, optional): List of column keys to order columns. Defaults to None.
        header_map (dict, optional): Mapping of column keys to header names. Defaults to None.
    """
    if not data:
        raise ValueError("No data to export.")

    if column_order is None:
        column_order = list(data[0].keys())

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to get active worksheet from Workbook.")

    # Write header
    headers = [header_map.get(col, col) if header_map else col for col in column_order]
    if hasattr(ws, "append") and callable(ws.append):
        ws.append(headers)
    else:
        raise RuntimeError("Worksheet object does not support 'append' method.")

    # Write data rows
    for row in data:
        ws.append([row.get(col, "") for col in column_order])

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
    wb.save(output_file_path)


def format_excel_columns(
    output_file_path: str, column_widths: Mapping[str, float]
) -> None:
    """
    Format Excel columns by setting widths and header styles.

    Args:
        output_file_path (str): Path to the Excel file to format.
        column_widths (Mapping[str, float]): Mapping of column letters to widths.
    """
    # type: ignore[import-untyped]
    wb = load_workbook(output_file_path)
    ws = wb.active
    assert ws is not None, "Active worksheet is None."
    ws = ws  # type: ignore[assignment]

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    font_color = Font(color="FFFFFF")

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    for column_letter, width in column_widths.items():
        header_cell = ws[f"{column_letter}1"]
        header_cell.fill = header_fill
        header_cell.font = font_color

    wb.save(output_file_path)


def format_excel_column_range(
    output_file_path: str,
    start_column: str,
    end_column: str,
    column_color: str = "00FF00",
    font_color: str = "FFFFFF",
) -> None:
    """
    Format a range of Excel columns with header color and font.

    Args:
        output_file_path (str): Path to the Excel file to format.
        start_column (str): Starting column letter (e.g., 'A').
        end_column (str): Ending column letter (e.g., 'D').
        column_color (str, optional): Header fill color. Defaults to '00FF00'.
        font_color (str, optional): Header font color. Defaults to 'FFFFFF'.
    """
    # type: ignore[import-untyped]
    wb = load_workbook(output_file_path)
    ws = wb.active
    assert ws is not None, "Active worksheet is None."
    ws = ws  # type: ignore[assignment]

    header_fill = PatternFill(
        start_color=column_color, end_color=column_color, fill_type="solid"
    )
    header_font = Font(color=font_color)

    for col_idx in range(
        column_index_from_string(start_column), column_index_from_string(end_column) + 1
    ):
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}1"]
        cell.fill = header_fill
        cell.font = header_font

    wb.save(output_file_path)


def add_filtering_to_all_columns(output_file_path: str) -> None:
    """
    Add auto-filtering and freeze header row in Excel file.

    Args:
        output_file_path (str): Path to the Excel file to modify.
    """
    # type: ignore[import-untyped]
    wb = load_workbook(output_file_path)
    ws = wb.active
    assert ws is not None, "Active worksheet is None."
    ws = ws  # type: ignore[assignment]

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(output_file_path)


def highlight_false_or_empty_values(
    output_file_path: str, columns_to_check: List[str]
) -> None:
    """
    Highlight cells with False, None, or empty string values in specified columns.

    Args:
        output_file_path (str): Path to the Excel file to modify.
        columns_to_check (list[str]): List of column letters to check.
    """
    # type: ignore[import-untyped]
    wb = load_workbook(output_file_path)
    ws = wb.active
    assert ws is not None, "Active worksheet is None."
    ws = ws  # type: ignore[assignment]

    red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    for col_letter in columns_to_check:
        for row in range(2, ws.max_row + 1):
            cell_value = ws[col_letter + str(row)].value
            if cell_value is False or cell_value is None or cell_value == "":
                ws[col_letter + str(row)].fill = red_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(output_file_path)


def validate_data(output_file_path: str, column_pairs: list[tuple[str, str]]) -> None:
    """
    Validate data by comparing pairs of columns and highlight mismatches.

    Args:
        output_file_path (str): Path to the Excel file to validate.
        column_pairs (list[tuple[str, str]]): List of (original_col, compare_col) pairs.
    """
    # type: ignore[import-untyped]
    wb = load_workbook(output_file_path)
    ws = wb.active
    assert ws is not None, "Active worksheet is None."
    ws = ws  # type: ignore[assignment]

    red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    for col_pair in column_pairs:
        original_col, compare_col = col_pair
        for row in range(2, ws.max_row + 1):
            original_value = ws[original_col + str(row)].value
            compare_value = ws[compare_col + str(row)].value
            if original_value != compare_value:
                ws[compare_col + str(row)].fill = red_fill

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_file_path)
